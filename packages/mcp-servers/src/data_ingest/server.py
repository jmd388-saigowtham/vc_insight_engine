"""Data ingestion and profiling — reads CSV/XLSX, profiles columns, samples rows."""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from pydantic import BaseModel, Field

from shared.python.schemas import ColumnProfile

# ---------------------------------------------------------------------------
# Path validation
# ---------------------------------------------------------------------------

_ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}


def _validate_path(file_path: str) -> Path:
    p = Path(file_path).resolve()
    if not p.exists():
        raise FileNotFoundError(f"File not found: {p}")
    if p.suffix.lower() not in _ALLOWED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {p.suffix}")
    return p


def _read_dataframe(
    path: Path, nrows: int | None = None, sheet_name: str | int = 0,
) -> pd.DataFrame:
    """Read a CSV or XLSX file into a DataFrame."""
    ext = path.suffix.lower()
    if ext == ".csv":
        file_size = path.stat().st_size
        if file_size > 100 * 1024 * 1024 and nrows is not None:
            # Large file — use chunked reading
            chunks: list[pd.DataFrame] = []
            rows_read = 0
            for chunk in pd.read_csv(path, chunksize=50_000):
                remaining = nrows - rows_read
                if remaining <= 0:
                    break
                chunks.append(chunk.head(remaining))
                rows_read += len(chunks[-1])
            return pd.concat(chunks, ignore_index=True) if chunks else pd.DataFrame()
        return pd.read_csv(path, nrows=nrows)
    else:
        return pd.read_excel(path, engine="openpyxl", nrows=nrows, sheet_name=sheet_name)


# ---------------------------------------------------------------------------
# Pydantic I/O models
# ---------------------------------------------------------------------------


class ProfileInput(BaseModel):
    file_path: str
    sample_size: int = 100_000
    sheet_name: str | int = 0


class ProfileOutput(BaseModel):
    columns: list[ColumnProfile]
    row_count: int
    file_size_bytes: int


class SampleInput(BaseModel):
    file_path: str
    n: int = 100
    sheet_name: str | int = 0


class SampleOutput(BaseModel):
    rows: list[dict[str, Any]]
    columns: list[str]


class RowCountInput(BaseModel):
    file_path: str


class RowCountOutput(BaseModel):
    count: int


class SheetInfo(BaseModel):
    name: str
    index: int


class ListSheetsInput(BaseModel):
    file_path: str


class ListSheetsOutput(BaseModel):
    sheets: list[SheetInfo]
    is_multi_sheet: bool


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def _safe_str(value: Any) -> str | None:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    return str(value)


def list_sheets(input: ListSheetsInput) -> ListSheetsOutput:
    """List all sheets in an xlsx file. Returns single sheet for CSV."""
    path = _validate_path(input.file_path)
    if path.suffix.lower() == ".csv":
        return ListSheetsOutput(
            sheets=[SheetInfo(name="Sheet1", index=0)],
            is_multi_sheet=False,
        )
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = [SheetInfo(name=name, index=i) for i, name in enumerate(wb.sheetnames)]
    wb.close()
    return ListSheetsOutput(
        sheets=sheets,
        is_multi_sheet=len(sheets) > 1,
    )


def profile(input: ProfileInput) -> ProfileOutput:
    """Profile all columns in the given file."""
    path = _validate_path(input.file_path)
    file_size = path.stat().st_size

    df = _read_dataframe(path, nrows=input.sample_size, sheet_name=input.sheet_name)
    total_rows = len(df)

    profiles: list[ColumnProfile] = []
    for col in df.columns:
        series = df[col]
        null_count = int(series.isna().sum())
        null_pct = round(null_count / total_rows * 100, 2) if total_rows > 0 else 0.0
        unique_count = int(series.nunique())

        min_val: str | None = None
        max_val: str | None = None
        mean_val: float | None = None

        if pd.api.types.is_numeric_dtype(series):
            min_val = _safe_str(series.min())
            max_val = _safe_str(series.max())
            mean_val = float(series.mean()) if not series.isna().all() else None
        elif pd.api.types.is_datetime64_any_dtype(series):
            min_val = _safe_str(series.min())
            max_val = _safe_str(series.max())

        sample_vals = series.dropna().unique()[:5].tolist()

        profiles.append(
            ColumnProfile(
                column_name=str(col),
                dtype=str(series.dtype),
                null_count=null_count,
                null_pct=null_pct,
                unique_count=unique_count,
                min_value=min_val,
                max_value=max_val,
                mean_value=mean_val,
                sample_values=sample_vals,
            )
        )

    return ProfileOutput(
        columns=profiles,
        row_count=total_rows,
        file_size_bytes=file_size,
    )


def sample(input: SampleInput) -> SampleOutput:
    """Return the first *n* rows as a list of dicts."""
    path = _validate_path(input.file_path)
    df = _read_dataframe(path, nrows=input.n, sheet_name=input.sheet_name)
    # Replace NaN with None for JSON serialisability
    df = df.where(pd.notnull(df), None)
    return SampleOutput(
        rows=df.to_dict(orient="records"),
        columns=[str(c) for c in df.columns],
    )


def row_count(input: RowCountInput) -> RowCountOutput:
    """Fast row count without loading the full file into memory."""
    path = _validate_path(input.file_path)
    ext = path.suffix.lower()

    if ext == ".csv":
        # Count newlines without loading data
        count = 0
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                count += chunk.count(b"\n")
        # Subtract 1 for header row
        count = max(count - 1, 0)
        return RowCountOutput(count=count)
    else:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        count = ws.max_row - 1 if ws.max_row else 0  # subtract header
        wb.close()
        return RowCountOutput(count=max(count, 0))
