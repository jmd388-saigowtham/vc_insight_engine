from __future__ import annotations

import uuid
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_db_session
from app.schemas.profile import ColumnProfileResponse, ProfileSummary
from app.services.profiling_service import ProfilingService

router = APIRouter()


def _get_service(db: AsyncSession = Depends(get_db_session)) -> ProfilingService:
    return ProfilingService(db)


@router.get("/files/{file_id}/profile", response_model=list[ColumnProfileResponse])
async def get_file_profile(
    file_id: uuid.UUID,
    service: ProfilingService = Depends(_get_service),
) -> list[ColumnProfileResponse]:
    profiles = await service.get_profiles(file_id)
    return [ColumnProfileResponse.model_validate(p) for p in profiles]


class DescriptionUpdate(BaseModel):
    description: str


@router.patch("/columns/{column_id}/description", response_model=ColumnProfileResponse)
async def update_column_description(
    column_id: uuid.UUID,
    data: DescriptionUpdate,
    service: ProfilingService = Depends(_get_service),
) -> ColumnProfileResponse:
    profile = await service.update_description(column_id, data.description)
    if profile is None:
        raise HTTPException(status_code=404, detail="Column profile not found")
    return ColumnProfileResponse.model_validate(profile)


class SheetInfoResponse(BaseModel):
    name: str
    index: int


class ListSheetsResponse(BaseModel):
    sheets: list[SheetInfoResponse] = Field(default_factory=list)
    is_multi_sheet: bool = False


@router.get("/files/{file_id}/sheets", response_model=ListSheetsResponse)
async def list_sheets(
    file_id: uuid.UUID,
    db: AsyncSession = Depends(get_db_session),
) -> ListSheetsResponse:
    """List sheets in an uploaded xlsx file."""
    from app.models.uploaded_file import UploadedFile

    file_obj = await db.get(UploadedFile, file_id)
    if file_obj is None:
        raise HTTPException(status_code=404, detail="File not found")

    storage_path = file_obj.storage_path
    if not storage_path:
        raise HTTPException(status_code=404, detail="File path not available")

    p = Path(storage_path)
    if not p.exists():
        from app.config import settings
        p = Path(settings.upload_dir) / storage_path
    if not p.exists():
        raise HTTPException(status_code=404, detail="File not found on disk")

    ext = p.suffix.lower()
    if ext == ".csv":
        return ListSheetsResponse(
            sheets=[SheetInfoResponse(name="Sheet1", index=0)],
            is_multi_sheet=False,
        )

    try:
        from openpyxl import load_workbook
        wb = load_workbook(p, read_only=True, data_only=True)
        sheets = [
            SheetInfoResponse(name=name, index=i)
            for i, name in enumerate(wb.sheetnames)
        ]
        wb.close()
        return ListSheetsResponse(
            sheets=sheets,
            is_multi_sheet=len(sheets) > 1,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read sheets: {e}")


@router.get("/sessions/{session_id}/tables", response_model=list[ProfileSummary])
async def get_session_tables(
    session_id: uuid.UUID,
    service: ProfilingService = Depends(_get_service),
) -> list[ProfileSummary]:
    tables = await service.get_session_tables(session_id)
    return [
        ProfileSummary(
            file_id=t["file_id"],
            filename=t["filename"],
            row_count=t["row_count"],
            column_count=t["column_count"],
            columns=[ColumnProfileResponse.model_validate(c) for c in t["columns"]],
        )
        for t in tables
    ]
